import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import io
import base64
from datetime import datetime
import warnings
import time
from scipy.optimize import curve_fit
from scipy.signal import savgol_filter
from sklearn.metrics import r2_score
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
warnings.filterwarnings('ignore')

# Configure matplotlib for better display
plt.style.use('default')

# Curve fitting functions
def exponential_rise(t, T_inf, T_0, tau):
    """Exponential temperature rise model: T = T_inf + (T_0 - T_inf) * exp(-t/tau)"""
    return T_inf + (T_0 - T_inf) * np.exp(-t / tau)

def linear_model(t, a, b):
    """Linear model: T = a*t + b"""
    return a * t + b

def polynomial_model(t, a, b, c):
    """Quadratic polynomial model: T = a*t^2 + b*t + c"""
    return a * t**2 + b * t + c

def logarithmic_model(t, a, b):
    """Logarithmic model: T = a*ln(t+1) + b"""
    return a * np.log(t + 1) + b

def fit_temperature_curve(time_data, temp_data, model_type='exponential'):
    """Fit curve to temperature data and return parameters and R²"""
    try:
        if model_type == 'exponential':
            # Initial guess for exponential model
            T_0 = temp_data[0] if len(temp_data) > 0 else 20
            T_inf = temp_data[-1] if len(temp_data) > 0 else 100
            tau = np.max(time_data) / 3 if len(time_data) > 0 else 100
            
            popt, pcov = curve_fit(exponential_rise, time_data, temp_data, 
                                 p0=[T_inf, T_0, tau], maxfev=5000)
            
            # Calculate R²
            y_pred = exponential_rise(time_data, *popt)
            r2 = r2_score(temp_data, y_pred)
            
            return popt, r2, exponential_rise
            
        elif model_type == 'linear':
            popt, pcov = curve_fit(linear_model, time_data, temp_data)
            y_pred = linear_model(time_data, *popt)
            r2 = r2_score(temp_data, y_pred)
            return popt, r2, linear_model
            
        elif model_type == 'polynomial':
            popt, pcov = curve_fit(polynomial_model, time_data, temp_data)
            y_pred = polynomial_model(time_data, *popt)
            r2 = r2_score(temp_data, y_pred)
            return popt, r2, polynomial_model
            
        elif model_type == 'logarithmic':
            popt, pcov = curve_fit(logarithmic_model, time_data, temp_data)
            y_pred = logarithmic_model(time_data, *popt)
            r2 = r2_score(temp_data, y_pred)
            return popt, r2, logarithmic_model
            
    except Exception as e:
        st.warning(f"Curve fitting failed for {model_type} model: {str(e)}")
        return None, 0, None

def generate_future_projection(params, model_func, time_data, extension_minutes):
    """Generate future temperature projection"""
    max_time = np.max(time_data)
    future_time = np.linspace(max_time, max_time + extension_minutes * 60, 100)
    future_temps = model_func(future_time, *params)
    return future_time, future_temps

def export_trimmed_data_to_excel(time_hours, temp_data, sample_interval_sec, plot_data, original_filename="temperature_data", report_meta=None):
    """
    Export trimmed temperature data to a formatted Excel file with embedded plot.
    
    Args:
        time_hours: Time array in hours (trimmed)
        temp_data: Dictionary of temperature data by channel (trimmed)
        sample_interval_sec: Sample interval in seconds
        plot_data: Dictionary containing plot configuration and results
        original_filename: Base filename for the export
        report_meta: Optional metadata for the report
    
    Returns:
        BytesIO object containing the Excel file
    """
    report_meta = report_meta or {}
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trimmed Temperature Data"
    
    # Create the data DataFrame
    export_df = pd.DataFrame()
    export_df['Time (hours)'] = time_hours
    export_df['Time (minutes)'] = time_hours * 60
    export_df['Sample #'] = range(1, len(time_hours) + 1)
    
    # Add temperature columns
    for channel_name, temps in temp_data.items():
        export_df[f'{channel_name} (°C)'] = temps
    
    # Write data to worksheet starting from row 8 to leave space for headers and metadata
    header_row = 8
    
    # Add title and metadata
    ws['A1'] = f"Temperature Data Export - {original_filename}"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = f"Exported on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(size=10, italic=True)
    # Optional test metadata
    meta_rows = [
        (3, 'Test Name', report_meta.get('test_name')),
        (4, 'Test Description', report_meta.get('test_description')),
        (5, 'Test Date', report_meta.get('test_date')),
        (6, 'Test Person', report_meta.get('test_person')),
    ]
    for row_idx, label, value in meta_rows:
        if value:
            ws[f'A{row_idx}'] = f"{label}: {value}"
            ws[f'A{row_idx}'].font = Font(size=10)
    
    # Write column headers with formatting
    for col_idx, col_name in enumerate(export_df.columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
            left=Side(border_style="thin"),
            right=Side(border_style="thin")
        )
    
    # Write data rows
    for row_idx, row_data in enumerate(dataframe_to_rows(export_df, index=False, header=False), start=header_row + 1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # Format temperature columns with 2 decimal places
            if col_idx > 3:  # Temperature columns
                cell.number_format = '0.00'
            elif col_idx == 2:  # Time minutes column
                cell.number_format = '0.0'
            elif col_idx == 1:  # Time hours column
                cell.number_format = '0.000'
            
            # Add borders
            cell.border = Border(
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
                left=Side(border_style="thin"),
                right=Side(border_style="thin")
            )
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)  # Cap at 20 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Generate and embed plot
    try:
        # Create the plot
        fig, ax = plt.subplots(figsize=(12, 8))
        
        # Apply filtering if enabled (match main plot behavior)
        filtered_data = {}
        filter_params = plot_data.get('filter_params', {}) or {}
        
        for channel_name, temps in temp_data.items():
            # Normalize to a clean series
            series = pd.Series(temps).dropna()
            if len(series) < 1:
                filtered_data[channel_name] = series  # keep empty; will be skipped in plotting
                continue
            
            if filter_params.get('enabled', False):
                try:
                    if filter_params.get('type') == 'moving_average':
                        window = min(filter_params.get('window', 5), len(series))
                        filt = series.rolling(window=window, center=True).mean()
                        filt = filt.bfill().ffill()
                        filtered_data[channel_name] = filt
                    elif filter_params.get('type') == 'savgol':
                        window = min(filter_params.get('window', 5), len(series))
                        if window % 2 == 0:
                            window -= 1
                        if window < 5:
                            window = min(5, len(series)) if len(series) >= 5 else len(series) | 1
                        polyorder = min(filter_params.get('polyorder', 2), max(1, window - 1))
                        filtered_values = savgol_filter(series.values, window_length=window, polyorder=polyorder)
                        filtered_data[channel_name] = pd.Series(filtered_values, index=series.index)
                    elif filter_params.get('type') == 'exponential':
                        alpha = filter_params.get('alpha', 0.3)
                        filtered_data[channel_name] = series.ewm(alpha=alpha).mean()
                    else:
                        filtered_data[channel_name] = series
                except Exception:
                    filtered_data[channel_name] = series
            else:
                filtered_data[channel_name] = series
        
        # Plot the data (align x/y lengths and skip empties)
        for channel_name, series in filtered_data.items():
            series_clean = pd.Series(series).dropna()
            if len(series_clean) == 0:
                continue
            ax.plot(time_hours[:len(series_clean)], series_clean.values, label=channel_name, linewidth=1.5)
        
        # Customize plot
        ax.set_xlabel('Time (hours)', fontsize=12)
        ax.set_ylabel('Temperature (°C)', fontsize=12)
        ax.set_title(f'Temperature Data - {original_filename}\n(Trimmed Dataset)', fontsize=14, fontweight='bold')
        ax.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
        ax.grid(True, alpha=0.3)
        
        # Add filter info to title if filtering was applied
        if filter_params.get('enabled', False):
            ftype = filter_params.get('type', '').replace('_', ' ').title()
            extra = []
            if ftype == 'Moving Average':
                extra.append(f"window={filter_params.get('window')}")
            elif ftype == 'Savgol':
                extra.append(f"window={filter_params.get('window')}, poly={filter_params.get('polyorder')}")
            elif ftype == 'Exponential':
                extra.append(f"α={filter_params.get('alpha')}")
            info = ", ".join([e for e in extra if e])
            filter_info = f"\nSmoothed: {ftype}{(' (' + info + ')') if info else ''}"
            current_title = ax.get_title()
            ax.set_title(current_title + filter_info, fontsize=12)
        
        plt.tight_layout()
        
        # Save plot to BytesIO
        plot_buffer = io.BytesIO()
        fig.savefig(plot_buffer, format='png', dpi=150, bbox_inches='tight')
        plot_buffer.seek(0)
        plt.close(fig)
        
        # Create new worksheet for the plot
        ws_plot = wb.create_sheet(title="Temperature Plot")
        
        # Insert the plot image
        img = OpenpyxlImage(plot_buffer)
        img.width = 800  # Adjust size as needed
        img.height = 600
        ws_plot.add_image(img, 'A1')
        
        # Add plot description
        ws_plot['A35'] = f"Temperature plot generated from trimmed dataset"
        ws_plot['A36'] = f"Data points: {len(time_hours)}"
        if len(time_hours) > 0:
            ws_plot['A37'] = f"Time range: {time_hours[0]:.2f} - {time_hours[-1]:.2f} hours"
        else:
            ws_plot['A37'] = "Time range: N/A"
        ws_plot['A38'] = f"Sample interval: {sample_interval_sec} seconds"
        
        # Format plot description
        for row in range(35, 39):
            ws_plot[f'A{row}'].font = Font(size=10)
        
    except Exception as e:
        # If plot generation fails, add error message to a separate sheet
        ws_error = wb.create_sheet(title="Plot Error")
        ws_error['A1'] = f"Plot generation failed: {str(e)}"
        ws_error['A1'].font = Font(color="FF0000")
    
    # Add Stability Analysis sheet if available
    try:
        stability = plot_data.get('stability_analysis') if isinstance(plot_data, dict) else None
        if stability:
            ws_stab = wb.create_sheet(title="Stability Analysis")
            # Headers
            headers = [
                'Channel', 'Status', 'Sample 1 (°C)', 'Sample 2 (°C)', 'Sample 3 (°C)',
                'Ref 1', 'Ref 2', 'Ref 3', 'Start Time (h)', 'End Time (h)',
                'Duration (h)', 'Avg Temp (°C)', 'Max Deviation (°C)', 'Reason'
            ]
            for c, h in enumerate(headers, start=1):
                cell = ws_stab.cell(row=1, column=c, value=h)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(
                    top=Side(border_style="thin"), bottom=Side(border_style="thin"),
                    left=Side(border_style="thin"), right=Side(border_style="thin")
                )

            r = 2
            for channel, result in stability.items():
                stable_flag = result.get('stable', False)
                status = 'Stable' if stable_flag else 'Unstable'
                points = result.get('stability_points') or []
                def fmt_point(i):
                    if i < len(points):
                        p = points[i]
                        temp = p.get('temp')
                        ref = p.get('san') if p.get('san') not in (None, '') else p.get('index')
                        return temp, ref
                    return None, None
                t1, r1 = fmt_point(0)
                t2, r2 = fmt_point(1)
                t3, r3 = fmt_point(2)
                values = [
                    channel,
                    status,
                    (round(float(t1), 2) if t1 is not None else None),
                    (round(float(t2), 2) if t2 is not None else None),
                    (round(float(t3), 2) if t3 is not None else None),
                    (str(r1) if r1 is not None else None),
                    (str(r2) if r2 is not None else None),
                    (str(r3) if r3 is not None else None),
                    (round(float(result.get('stability_start_time')), 3) if result.get('stability_start_time') is not None else None),
                    (round(float(result.get('stability_end_time')), 3) if result.get('stability_end_time') is not None else None),
                    (round(float(result.get('stability_duration')), 3) if result.get('stability_duration') is not None else None),
                    (round(float(result.get('average_temp')), 2) if result.get('average_temp') is not None else None),
                    (round(float(result.get('max_deviation')), 2) if result.get('max_deviation') is not None else None),
                    result.get('reason')
                ]

                for c, v in enumerate(values, start=1):
                    cell = ws_stab.cell(row=r, column=c, value=v)
                    if c in (3, 4, 5, 12, 13):
                        cell.number_format = '0.00'
                    elif c in (9, 10, 11):
                        cell.number_format = '0.000'
                    cell.border = Border(
                        top=Side(border_style="thin"), bottom=Side(border_style="thin"),
                        left=Side(border_style="thin"), right=Side(border_style="thin")
                    )
                r += 1

            # Auto-size columns
            for column in ws_stab.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws_stab.column_dimensions[column_letter].width = min(max_length + 2, 28)
    except Exception:
        # Don't fail export if stability formatting hits an edge case
        pass

    # Save workbook to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

def analyze_temperature_stability(time_hours, temp_data, sample_interval_sec, stability_threshold=1.0, interval_minutes=15, row_offset=0, san_series=None):
    """
    New stability analysis:
    - Identify last sample (T), T-30, T-60, T-90 (row offsets)
    - Compute differences: d1 = T - T-30, d2 = T-30 - T-60, d3 = T-60 - T-90
    - Unstable if any |d| > stability_threshold
    Returns dictionary keyed by channel with required fields for UI/export.
    """
    results = {}

    if len(time_hours) == 0:
        return results

    for channel, temps in temp_data.items():
        temps_clean = temps.dropna()
        n = len(temps_clean)

        # Need last, last-30, last-60, last-90
        required = 91  # last index -90 must be >= 0
        if n < required:
            results[channel] = {
                'stable': False,
                'reason': f'Insufficient data (need {required} samples, have {n})',
                'stability_points': [],
                'differences': {},
                'max_deviation': None,
                'stability_duration': None,
                'stability_start_time': None,
                'stability_end_time': None,
                'average_temp': None
            }
            continue

        # Indices in cleaned series
        idx0 = n - 1
        idx30 = n - 1 - 30
        idx60 = n - 1 - 60
        idx90 = n - 1 - 90

        T0 = float(temps_clean.iloc[idx0])
        T30 = float(temps_clean.iloc[idx30])
        T60 = float(temps_clean.iloc[idx60])
        T90 = float(temps_clean.iloc[idx90])

        d1 = T0 - T30
        d2 = T30 - T60
        d3 = T60 - T90

        abs_diffs = [abs(d1), abs(d2), abs(d3)]
        max_dev = max(abs_diffs)

        # Determine stability
        unstable_flags = [abs(d1) > stability_threshold, abs(d2) > stability_threshold, abs(d3) > stability_threshold]
        is_stable = not any(unstable_flags)

        # Map to original indices for SAN/row mapping
        orig_idx0 = int(temps_clean.index[idx0])
        orig_idx30 = int(temps_clean.index[idx30])
        orig_idx60 = int(temps_clean.index[idx60])
        san0 = (None if san_series is None or orig_idx0 >= len(san_series) else san_series.iloc[orig_idx0])
        san30 = (None if san_series is None or orig_idx30 >= len(san_series) else san_series.iloc[orig_idx30])
        san60 = (None if san_series is None or orig_idx60 >= len(san_series) else san_series.iloc[orig_idx60])

        stability_points = [
            {'time_hours': time_hours[idx0] if idx0 < len(time_hours) else None, 'temp': T0, 'row': orig_idx0 + row_offset, 'index': orig_idx0 + row_offset, 'san': san0},
            {'time_hours': time_hours[idx30] if idx30 < len(time_hours) else None, 'temp': T30, 'row': orig_idx30 + row_offset, 'index': orig_idx30 + row_offset, 'san': san30},
            {'time_hours': time_hours[idx60] if idx60 < len(time_hours) else None, 'temp': T60, 'row': orig_idx60 + row_offset, 'index': orig_idx60 + row_offset, 'san': san60},
        ]

        # Build reason text
        if is_stable:
            reason = 'Temperature stable: all 30-sample changes ≤ {:.1f}°C'.format(stability_threshold)
        else:
            reasons = []
            if unstable_flags[0]:
                reasons.append(f"Difference between T and T-30 was {abs(d1):.1f}°C")
            if unstable_flags[1]:
                reasons.append(f"Difference between T-30 and T-60 was {abs(d2):.1f}°C")
            if unstable_flags[2]:
                reasons.append(f"Difference between T-60 and T-90 was {abs(d3):.1f}°C")
            reason = 'Unstable: ' + '; '.join(reasons)

        results[channel] = {
            'stable': is_stable,
            'reason': reason,
            'stability_points': stability_points,
            'differences': {
                'd1 (T vs T-30)': round(d1, 3),
                'd2 (T-30 vs T-60)': round(d2, 3),
                'd3 (T-60 vs T-90)': round(d3, 3)
            },
            'max_deviation': max_dev,
            'stability_duration': None,
            'stability_start_time': None,
            'stability_end_time': None,
            'average_temp': float(np.mean([p['temp'] for p in stability_points])) if stability_points else None
        }

    return results



def generate_final_data_table(time_hours, temp_data, sample_interval_sec, temp_rating=55.0):
    """Generate final data table with worst-case analysis"""
    
    # Find the time when hottest temperature occurs across all channels
    max_temp = -999
    max_temp_time = 0
    max_temp_sample = 0
    hottest_channel = ""
    
    for channel, temps in temp_data.items():
        temp_series = temps.dropna()
        if not temp_series.empty:
            channel_max = temp_series.max()
            if channel_max > max_temp:
                max_temp = channel_max
                max_temp_idx = temp_series.idxmax()
                max_temp_time = time_hours[max_temp_idx] if max_temp_idx < len(time_hours) else 0
                max_temp_sample = max_temp_idx
                hottest_channel = channel
    
    # Get temperatures of all channels at the hottest time
    temps_at_hottest_time = {}
    for channel, temps in temp_data.items():
        temp_series = temps.dropna()
        if max_temp_sample < len(temp_series):
            temps_at_hottest_time[channel] = temp_series.iloc[max_temp_sample]
        else:
            temps_at_hottest_time[channel] = np.nan
    
    # Identify ambient channel (assume it's the coolest at hottest time)
    valid_temps = {k: v for k, v in temps_at_hottest_time.items() if not np.isnan(v)}
    if valid_temps:
        ambient_channel = min(valid_temps, key=valid_temps.get)
        ambient_temp = valid_temps[ambient_channel]
    else:
        ambient_channel = ""
        ambient_temp = 20.0  # Default ambient
    
    # Calculate temperature rises and worst-case temperatures
    final_data = []
    for channel, temp_at_hottest in temps_at_hottest_time.items():
        if not np.isnan(temp_at_hottest):
            temp_rise = temp_at_hottest - ambient_temp
            worst_case_temp = temp_rise + temp_rating
            
            final_data.append({
                'Channel': channel,
                'Temp at Hottest Time (°C)': round(temp_at_hottest, 2),
                'Temperature Rise (°C)': round(temp_rise, 2),
                'Worst Case Temp (°C)': round(worst_case_temp, 2),
                'Is Hottest': channel == hottest_channel,
                'Is Ambient': channel == ambient_channel
            })
    
    return {
        'table_data': final_data,
        'hottest_time_hours': max_temp_time,
        'hottest_sample': max_temp_sample,
        'hottest_channel': hottest_channel,
        'max_temperature': max_temp,
        'ambient_channel': ambient_channel,
        'ambient_temp': ambient_temp,
        'temp_rating': temp_rating
    }

## Smart detection removed per user request

def apply_global_trimming(time_hours, temp_data):
    """Apply global trimming settings to time and temperature data"""
    if not st.session_state.get('global_trim_enabled', False):
        return time_hours, temp_data
    
    # Get trim settings
    start_time = st.session_state.get('global_start_time')
    end_time = st.session_state.get('global_end_time')
    preserve_original_time = st.session_state.get('global_preserve_original_time', False)
    
    # Calculate trim indices
    start_idx = 0
    end_idx = len(time_hours) - 1
    
    if start_time is not None:
        start_idx = np.argmin(np.abs(time_hours - start_time))
    if end_time is not None:
        end_idx = np.argmin(np.abs(time_hours - end_time))
    
    # Apply trimming
    trimmed_time = time_hours[start_idx:end_idx+1]
    
    # Reset time to start from 0 if preserve_original_time is False
    # Use the actual first trimmed time to avoid floating mismatch with requested start_time
    if start_time is not None and not preserve_original_time and len(trimmed_time) > 0:
        trimmed_time = trimmed_time - trimmed_time[0]
    
    trimmed_temp_data = {}
    
    for channel, temps in temp_data.items():
        # Slice original series to keep alignment with time axis and avoid shifting due to dropna
        if start_idx < len(temps):
            actual_end_idx = min(end_idx, len(temps) - 1)
            trimmed_series = temps.iloc[start_idx:actual_end_idx+1]
            # Keep original NaNs; downstream can choose to handle them
            trimmed_temp_data[channel] = trimmed_series.reset_index(drop=True)
        else:
            trimmed_temp_data[channel] = pd.Series(dtype=float)
    
    return trimmed_time, trimmed_temp_data

def calculate_temperature_deltas(time_hours, temp_data, reference_time):
    """Calculate temperature change from reference time to final measurement"""
    delta_results = {}
    
    # Find the closest time point to reference_time
    if len(time_hours) == 0:
        return delta_results
    
    ref_idx = np.argmin(np.abs(time_hours - reference_time))
    actual_ref_time = time_hours[ref_idx]
    
    # Get the final time index for the data
    final_time_idx = len(time_hours) - 1
    actual_final_time = time_hours[final_time_idx]
    
    for channel, temps in temp_data.items():
        temp_series = temps.dropna()
        if len(temp_series) > ref_idx and not temp_series.empty:
            # Ensure we don't exceed available data
            final_idx = min(len(temp_series) - 1, final_time_idx)
            
            ref_temp = temp_series.iloc[ref_idx]
            final_temp = temp_series.iloc[final_idx]
            delta_temp = final_temp - ref_temp
            
            # Use the corresponding time from time_hours array
            channel_final_time = time_hours[final_idx]
            
            delta_results[channel] = {
                'reference_temp': ref_temp,
                'final_temp': final_temp,
                'delta_temp': delta_temp,
                'reference_time': actual_ref_time,
                'final_time': channel_final_time,
                'time_span': channel_final_time - actual_ref_time
            }
    
    return delta_results

    # comprehensive analysis report function removed per user request

def main():
    st.title("Temperature Data Logger Analysis")
    st.markdown("Upload Excel workbooks containing temperature logging data and generate publication-ready plots.")
    
    # Note: Database-backed configuration management removed per user request
    
    # File upload section
    st.header("📁 Data Upload")
    uploaded_file = st.file_uploader(
        "Choose an Excel file (.xlsx, .xls)", 
        type=['xlsx', 'xls'],
        help="Upload your Excel workbook containing temperature logging data"
    )
    
    if uploaded_file is not None:
        try:
            # Load the Excel file
            excel_file = pd.ExcelFile(uploaded_file)
            
            # Sheet selection
            sheet_names = excel_file.sheet_names
            selected_sheet = st.selectbox(
                "Select Sheet",
                options=sheet_names,
                index=0,
                help="Choose the sheet containing your temperature data"
            )
            
            # Load the selected sheet
            df = pd.read_excel(uploaded_file, sheet_name=selected_sheet, header=None)
            
            # Store file data in session state for saving configurations
            st.session_state.current_file_data = uploaded_file.getvalue()
            st.session_state.current_filename = uploaded_file.name
            st.session_state.current_sheet_name = selected_sheet
            
            st.success(f"✅ Loaded sheet '{selected_sheet}' with {df.shape[0]} rows and {df.shape[1]} columns")
            
            # Store uploaded df for later use
            st.session_state.uploaded_df = df
            
            # Initialize trimming variables with default values (must be early for global access)
            enable_start_trim = False
            enable_end_trim = False
            start_time_hours = 0.0
            end_time_hours = 100.0
            
            # Configuration section
            st.header("⚙️ Data Configuration")
            
            # Global configuration variables (accessible to all sections)
            col1, col2 = st.columns(2)
            
            with col1:
                st.subheader("Row Configuration")
                
                descriptor_row = st.number_input(
                    "Descriptor Row (0-based index)",
                    min_value=0,
                    max_value=df.shape[0]-1,
                    value=min(44, df.shape[0]-1),
                    help="Row containing column descriptions/names"
                )
                
                first_data_row = st.number_input(
                    "First Data Row (0-based index)",
                    min_value=int(descriptor_row)+1,
                    max_value=df.shape[0]-1,
                    value=min(int(descriptor_row)+1, df.shape[0]-1),
                    help="First row containing actual measurement data"
                )
                
                sample_interval_sec = st.number_input(
                    "Sample Interval (seconds)",
                    min_value=1,
                    value=30,
                    help="Time between successive measurements"
                )
                
                # Data preview controls with full range scrolling
                st.write("**Data Preview - Scrollable (0 to 100 rows):**")
                st.write("🟡 = Descriptor Row | 🟢 = First Data Row")
                
                # Row range controls
                col_preview1, col_preview2 = st.columns(2)
                
                with col_preview1:
                    preview_start = st.number_input(
                        "Start row for preview",
                        min_value=0,
                        max_value=max(0, df.shape[0]-1),
                        value=max(0, int(descriptor_row) - 10),
                        help="Starting row for data preview"
                    )
                
                with col_preview2:
                    preview_rows = st.number_input(
                        "Number of rows to show",
                        min_value=1,
                        max_value=100,
                        value=50,
                        help="How many rows to display in preview"
                    )
                
                start_row = int(preview_start)
                end_row = min(df.shape[0], start_row + int(preview_rows))
                
                st.write(f"**Showing rows {start_row} to {end_row-1}:**")
                
                preview_df = df.iloc[start_row:end_row].copy()
                # Convert all columns to string to avoid Arrow conversion issues
                preview_df = preview_df.astype(str)
                
                # Reset index to show actual row numbers
                preview_df.index = range(start_row, end_row)
                
                # Highlight descriptor and first data rows
                def highlight_rows(row):
                    if row.name == int(descriptor_row):
                        return ['background-color: #ffeb3b'] * len(row)  # Yellow for descriptor
                    elif row.name == int(first_data_row):
                        return ['background-color: #4caf50; color: white'] * len(row)  # Green for first data
                    return [''] * len(row)
                
                styled_df = preview_df.style.apply(highlight_rows, axis=1)
                st.dataframe(styled_df, use_container_width=True, height=400)  # Fixed height for scrolling
                

            
            # Preview trimmed data with overlay and trimmed-only view
            if (enable_start_trim or enable_end_trim) and 'graph_data' in st.session_state:
                col_trim_btn1, col_trim_btn2 = st.columns(2)
                
                with col_trim_btn1:
                    if st.button("📊 Show Trim Overlay", help="Show trim areas overlaid on the main graph"):
                        try:
                            # Use stored graph data
                            graph_data = st.session_state.graph_data
                            temp_data_stored = graph_data['temp_data']
                            time_hours_stored = graph_data['time_hours']
                            colors_stored = graph_data['colors']
                            
                            # Create overlay plot
                            fig, ax = plt.subplots(figsize=(12, 6))
                            
                            # Plot all original data
                            for i, (col, temps) in enumerate(temp_data_stored.items()):
                                temps_clean = temps.dropna()
                                if not temps_clean.empty:
                                    color_idx = i % 10
                                    ax.plot(time_hours_stored[:len(temps_clean)], temps_clean.values, 
                                           color=colors_stored[color_idx], label=col, linewidth=1.5, alpha=0.7)
                            
                            # Add shaded regions for trimmed areas
                            y_min, y_max = ax.get_ylim()
                            
                            if enable_start_trim:
                                # Shade area before start trim
                                ax.axvspan(0, start_time_hours, alpha=0.3, color='red', label='Trimmed Area (Start)')
                                ax.axvline(x=start_time_hours, color='green', linestyle='--', linewidth=2, label='Start Trim Point')
                            
                            if enable_end_trim:
                                # Shade area after end trim
                                max_time = time_hours_stored[-1]
                                ax.axvspan(end_time_hours, max_time, alpha=0.3, color='red', label='Trimmed Area (End)')
                                ax.axvline(x=end_time_hours, color='red', linestyle='--', linewidth=2, label='End Trim Point')
                            
                            ax.set_xlabel('Time (hours)')
                            ax.set_ylabel('Temperature (°C)')
                            ax.set_title('Data with Trim Areas Highlighted')
                            ax.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
                            ax.grid(True, alpha=0.3)
                            
                            plt.tight_layout()
                            st.pyplot(fig, use_container_width=True)
                            
                        except Exception as e:
                            st.error(f"Trim overlay failed: {str(e)}")
                
                with col_trim_btn2:
                    if st.button("✂️ Show Trimmed Data Only", help="Show only the selected data after trimming"):
                        try:
                            # Use stored graph data
                            graph_data = st.session_state.graph_data
                            temp_data_stored = graph_data['temp_data']
                            time_hours_stored = graph_data['time_hours']
                            colors_stored = graph_data['colors']
                            
                            # Create trimmed-only plot
                            fig, ax = plt.subplots(figsize=(12, 6))
                            
                            # Calculate trim indices
                            start_idx = 0
                            end_idx = len(time_hours_stored) - 1
                            
                            if enable_start_trim:
                                start_idx = np.argmin(np.abs(time_hours_stored - start_time_hours))
                            if enable_end_trim:
                                end_idx = np.argmin(np.abs(time_hours_stored - end_time_hours))
                            
                            # Plot only trimmed data
                            for i, (col, temps) in enumerate(temp_data_stored.items()):
                                temps_clean = temps.dropna()
                                if not temps_clean.empty:
                                    # Get trimmed data
                                    actual_end_idx = min(end_idx, len(temps_clean) - 1)
                                    if start_idx < len(temps_clean):
                                        trimmed_temps = temps_clean.iloc[start_idx:actual_end_idx+1]
                                        trimmed_time = time_hours_stored[start_idx:start_idx+len(trimmed_temps)]
                                        
                                        color_idx = i % 10
                                        ax.plot(trimmed_time, trimmed_temps.values, 
                                               color=colors_stored[color_idx], label=col, linewidth=1.5)
                            
                            ax.set_xlabel('Time (hours)')
                            ax.set_ylabel('Temperature (°C)')
                            ax.set_title('Analysis Data (After Trimming)')
                            ax.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
                            ax.grid(True, alpha=0.3)
                            
                            plt.tight_layout()
                            st.pyplot(fig, use_container_width=True)
                            
                        except Exception as e:
                            st.error(f"Trimmed data view failed: {str(e)}")
                
                # Show trim summary
                if enable_start_trim and enable_end_trim:
                    duration = end_time_hours - start_time_hours
                    retained_pct = (duration / st.session_state.graph_data['time_hours'][-1]) * 100
                    st.info(f"📊 Preview: {start_time_hours:.2f}h to {end_time_hours:.2f}h (Duration: {duration:.2f}h, {retained_pct:.1f}% of data)")
                elif enable_start_trim:
                    duration = st.session_state.graph_data['time_hours'][-1] - start_time_hours
                    retained_pct = (duration / st.session_state.graph_data['time_hours'][-1]) * 100
                    st.info(f"📊 Preview: Start from {start_time_hours:.2f}h ({retained_pct:.1f}% of data retained)")
                elif enable_end_trim:
                    retained_pct = (end_time_hours / st.session_state.graph_data['time_hours'][-1]) * 100
                    st.info(f"📊 Preview: End at {end_time_hours:.2f}h ({retained_pct:.1f}% of data retained)")
                    
            elif enable_start_trim or enable_end_trim:
                st.info("💡 Generate the data graph first to see trim options")
            
            # Apply trimming to all analysis button
            if enable_start_trim or enable_end_trim:
                st.markdown("---")
                apply_trim_col1, apply_trim_col2 = st.columns([2, 1])
                
                with apply_trim_col1:
                    if st.button("✅ Apply Trimming to All Analysis", 
                                help="Apply current trim settings to all subsequent graphs and analysis", 
                                type="primary"):
                        # Store trim settings in session state for global application
                        st.session_state.global_trim_enabled = True
                        st.session_state.global_start_trim = enable_start_trim
                        st.session_state.global_end_trim = enable_end_trim
                        st.session_state.global_start_time = start_time_hours if enable_start_trim else None
                        st.session_state.global_end_time = end_time_hours if enable_end_trim else None
                        st.session_state.global_preserve_original_time = preserve_original_time
                        
                        st.success("✅ Trimming settings applied! All future analysis will use only the selected data range.")
                        st.info("🔄 Any graphs you generate from now on will automatically exclude the trimmed regions.")
                
                with apply_trim_col2:
                    if st.button("🔄 Reset Trimming", help="Remove all trim settings"):
                        st.session_state.global_trim_enabled = False
                        st.session_state.global_start_trim = False
                        st.session_state.global_end_trim = False
                        st.session_state.global_start_time = None
                        st.session_state.global_end_time = None
                        st.success("🔄 Trimming reset. All data will be used for analysis.")
            
            # Show current global trim status
            if st.session_state.get('global_trim_enabled', False):
                trim_start = st.session_state.get('global_start_time')
                trim_end = st.session_state.get('global_end_time')
                
                if trim_start is not None and trim_end is not None:
                    duration = trim_end - trim_start
                    st.success(f"🎯 **Global Trimming Active**: Using data from {trim_start:.2f}h to {trim_end:.2f}h (Duration: {duration:.2f}h)")
                elif trim_start is not None:
                    st.success(f"🎯 **Global Trimming Active**: Using data from {trim_start:.2f}h onwards")
                elif trim_end is not None:
                    st.success(f"🎯 **Global Trimming Active**: Using data up to {trim_end:.2f}h")
            
            with col2:
                st.subheader("Column Configuration")
                
                id_column_pattern = st.text_input(
                    "ID Column Pattern",
                    value="name",
                    help="Substring to identify the sample counter column (case-insensitive)"
                )
                
                exclude_substring = st.text_input(
                    "Exclude Columns Containing",
                    value="alarm,scan",
                    help="Exclude columns whose descriptors contain these substrings (comma-separated, case-insensitive)"
                )
                
                legend_source = st.selectbox(
                    "Legend Label Source",
                    options=["descriptor", "header"],
                    index=0,
                    help="Use descriptor row or header row for legend labels"
                )
                
                # Store additional values in session state
                st.session_state.current_id_pattern = id_column_pattern
                st.session_state.current_exclude_substring = exclude_substring
                st.session_state.current_legend_source = legend_source
            
            # Store current values in session state (now accessible to all sections)
            st.session_state.current_descriptor_row = descriptor_row
            st.session_state.current_first_data_row = first_data_row
            st.session_state.current_sample_interval = sample_interval_sec
            
            # Plot configuration
            st.header("📊 Plot Configuration")
            
            col3, col4 = st.columns(2)
            
            with col3:
                tick_interval_min = st.number_input(
                    "X-axis Tick Interval (minutes)",
                    min_value=1,
                    value=30,
                    help="Spacing between x-axis time ticks"
                )
                
                colormap_name = st.selectbox(
                    "Colormap",
                    options=['turbo', 'viridis', 'plasma', 'inferno', 'magma', 'tab10', 'Set3', 'rainbow'],
                    index=0,
                    help="Matplotlib colormap for automatic color assignment"
                )
            
            with col4:
                export_format = st.selectbox(
                    "Export Format",
                    options=['PDF', 'JPEG', 'SVG'],
                    index=0,
                    help="Choose output format for the plot",
                    key="export_format_selectbox"
                )
                st.session_state['export_format_selection'] = export_format
                
                custom_colors_input = st.text_area(
                    "Custom Colors (optional)",
                    placeholder="Enter hex colors separated by commas, e.g., #FF0000, #00FF00, #0000FF",
                    help="Optional: Custom hex colors for specific styling"
                )
                
                # Store final values in session state
                st.session_state.current_tick_interval = tick_interval_min
                st.session_state.current_colormap = colormap_name
                
                # Process custom colors if provided
                custom_colors = []
                if custom_colors_input.strip():
                    try:
                        colors = [color.strip() for color in custom_colors_input.split(',')]
                        # Validate hex colors
                        for color in colors:
                            if not color.startswith('#') or len(color) != 7:
                                st.warning(f"Invalid hex color: {color}. Use format #RRGGBB")
                            else:
                                custom_colors.append(color)
                    except:
                        st.warning("Error parsing custom colors. Please use format: #FF0000, #00FF00, #0000FF")
            
            # Always compute and show the data graph
            st.header("📈 Data")
            try:
                temp_data_graph = {}
                descriptor_labels = {}
                descriptor_data = df.iloc[descriptor_row]
                # Find ID column using descriptor row and user pattern
                id_col_idx = None
                for idx, desc in enumerate(descriptor_data):
                    desc_str = str(desc).lower()
                    if id_column_pattern.lower() in desc_str:
                        id_col_idx = idx
                        break
                # Extract temperature columns using descriptor row
                exclude_patterns = [pattern.strip().lower() for pattern in exclude_substring.split(',') if pattern.strip()]
                for idx, desc in enumerate(descriptor_data):
                    if idx == id_col_idx:
                        continue
                    desc_str = str(desc).lower()
                    if any(pattern in desc_str for pattern in exclude_patterns):
                        continue
                    col_data = pd.to_numeric(df.iloc[first_data_row:, idx], errors='coerce')
                    if not col_data.dropna().empty:
                        temp_data_graph[str(desc)] = col_data
                        descriptor_labels[str(desc)] = str(desc)
                if temp_data_graph:
                    num_samples = len(temp_data_graph[list(temp_data_graph.keys())[0]])
                    time_hours_graph = np.arange(num_samples) * sample_interval_sec / 3600
                    colors = plt.cm.tab10(np.linspace(0, 1, min(10, len(temp_data_graph))))
                    # Determine highlight indices based on current trim values in session state
                    start_val = float(st.session_state.get('start_trim', 0.0))
                    end_default = float(time_hours_graph[-1]) if len(time_hours_graph) else 0.0
                    end_val = float(st.session_state.get('end_trim', end_default))
                    start_idx = int(np.argmin(np.abs(time_hours_graph - start_val))) if len(time_hours_graph) else 0
                    end_idx = int(np.argmin(np.abs(time_hours_graph - end_val))) if len(time_hours_graph) else 0
                    start_idx = max(0, start_idx)
                    end_idx = max(start_idx, min(end_idx, len(time_hours_graph) - 1))
                    fig, ax = plt.subplots(figsize=(12, 6))
                    # Plot all data
                    for i, (channel_name, temps) in enumerate(temp_data_graph.items()):
                        temps_clean = temps.dropna()
                        if not temps_clean.empty:
                            color_idx = i % 10
                            time_for_channel = time_hours_graph[:len(temps_clean)]
                            ax.plot(time_for_channel, temps_clean.values,
                                    color=colors[color_idx], label=channel_name, linewidth=1.5, alpha=0.9)
                    ax.set_xlabel('Time (hours)')
                    ax.set_ylabel('Temperature (°C)')
                    ax.set_title('Temperature Data')
                    ax.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
                    ax.grid(True, alpha=0.3)
                    # Highlight kept region if values indicate trimming
                    trim_active = (start_val > 0.0 + 1e-9) or (end_val < end_default - 1e-9)
                    if trim_active:
                        x_min = time_hours_graph[start_idx] if len(time_hours_graph) else 0.0
                        x_max = time_hours_graph[end_idx] if len(time_hours_graph) else 0.0
                        # Dim outside region
                        y_min, y_max = ax.get_ylim()
                        if start_idx > 0:
                            ax.axvspan(time_hours_graph[0], x_min, alpha=0.15, color='red', label='Trimmed')
                        if end_idx < len(time_hours_graph) - 1:
                            ax.axvspan(x_max, time_hours_graph[-1], alpha=0.15, color='red')
                        # Outline kept region
                        ax.axvspan(x_min, x_max, alpha=0.08, color='green', label='Kept')
                    plt.tight_layout()
                    st.pyplot(fig, use_container_width=True)
                    # Store for reuse
                    st.session_state.graph_data = {
                        'temp_data': temp_data_graph,
                        'time_hours': time_hours_graph,
                        'colors': colors
                    }
                else:
                    st.warning("No temperature data found with current configuration")
            except Exception as e:
                st.error(f"Graph generation failed: {str(e)}")
            
            # Manual trimming controls
            st.subheader("⚙️ Manual Trim Settings")
            st.info("Set the start/end times to trim data; highlight updates automatically.")
            
            # Time axis option
            preserve_original_time = st.checkbox(
                "Keep original time axis when trimming", 
                value=False,
                help="When enabled, trimmed data keeps the original time values. When disabled, time resets to start from 0."
            )
            
            col_trim1, col_trim2 = st.columns(2)
            
            with col_trim1:
                default_start = 0.0
                start_time_hours = st.number_input(
                    "Start Time (hours)",
                    min_value=0.0,
                    max_value=100.0,
                    value=st.session_state.get('start_trim', default_start),
                    step=0.1,
                    help="Remove all data before this time point",
                    key="start_trim"
                )
            
            with col_trim2:
                # Default end time to the dataset's end time
                default_end = (
                    st.session_state.graph_data['time_hours'][-1]
                    if 'graph_data' in st.session_state and 'time_hours' in st.session_state.graph_data and len(st.session_state.graph_data['time_hours']) > 0
                    else st.session_state.get('suggested_end', 24.0)
                )
                end_time_hours = st.number_input(
                    "End Time (hours)",
                    min_value=0.1,
                    max_value=200.0,
                    value=st.session_state.get('end_trim', default_end),
                    step=0.1,
                    help="Remove all data after this time point",
                    key="end_trim"
                )
            
            # Show trim summary (derive active flags from values)
                enable_start_trim = start_time_hours > 0.0 + 1e-9
                enable_end_trim = end_time_hours < st.session_state.graph_data['time_hours'][-1] if 'graph_data' in st.session_state else False
                if enable_start_trim and enable_end_trim:
                    duration = end_time_hours - start_time_hours
                    retained_pct = (duration / st.session_state.graph_data['time_hours'][-1]) * 100
                    st.info(f"📊 Preview: {start_time_hours:.2f}h to {end_time_hours:.2f}h (Duration: {duration:.2f}h, {retained_pct:.1f}% of data)")
                elif enable_start_trim:
                    duration = st.session_state.graph_data['time_hours'][-1] - start_time_hours
                    retained_pct = (duration / st.session_state.graph_data['time_hours'][-1]) * 100
                    st.info(f"📊 Preview: Start from {start_time_hours:.2f}h ({retained_pct:.1f}% of data retained)")
                elif enable_end_trim:
                    retained_pct = (end_time_hours / st.session_state.graph_data['time_hours'][-1]) * 100
                    st.info(f"📊 Preview: End at {end_time_hours:.2f}h ({retained_pct:.1f}% of data retained)")
            
            # Apply trimming to all analysis button (always shown)
            if True:
                st.markdown("---")
                apply_trim_col1, apply_trim_col2 = st.columns([2, 1])
                
                with apply_trim_col1:
                    if st.button("✅ Apply Trimming to All Analysis", 
                                help="Apply current trim settings to all subsequent graphs and analysis", 
                                type="primary"):
                        # Store trim settings in session state for global application
                        st.session_state.global_trim_enabled = True
                        st.session_state.global_start_trim = start_time_hours > 0.0 + 1e-9
                        st.session_state.global_end_trim = ('graph_data' in st.session_state and end_time_hours < st.session_state.graph_data['time_hours'][-1])
                        st.session_state.global_start_time = start_time_hours
                        st.session_state.global_end_time = end_time_hours
                        st.session_state.global_preserve_original_time = preserve_original_time
                        
                        st.success("✅ Trimming settings applied! All future analysis will use only the selected data range.")
                        st.info("🔄 Any graphs you generate from now on will automatically exclude the trimmed regions.")
                
                with apply_trim_col2:
                    if st.button("🔄 Reset Trimming", help="Remove all trim settings"):
                        st.session_state.global_trim_enabled = False
                        st.session_state.global_start_trim = False
                        st.session_state.global_end_trim = False
                        st.session_state.global_start_time = None
                        st.session_state.global_end_time = None
                        st.session_state.global_preserve_original_time = False
                        st.success("🔄 Trimming reset. All data will be used for analysis.")
            
            # Show current global trim status
            if st.session_state.get('global_trim_enabled', False):
                trim_start = st.session_state.get('global_start_time')
                trim_end = st.session_state.get('global_end_time')
                
                if trim_start is not None and trim_end is not None:
                    duration = trim_end - trim_start
                    st.success(f"🎯 **Global Trimming Active**: Using data from {trim_start:.2f}h to {trim_end:.2f}h (Duration: {duration:.2f}h)")
                elif trim_start is not None:
                    st.success(f"🎯 **Global Trimming Active**: Using data from {trim_start:.2f}h onwards")
                elif trim_end is not None:
                    st.success(f"🎯 **Global Trimming Active**: Using data up to {trim_end:.2f}h")
            
            # Curve Fitting Section
            st.header("📈 Curve Fitting & Future Projection")
            
            enable_curve_fitting = st.checkbox("Enable curve fitting and future projection", 
                                             help="Fit mathematical models to temperature rise data and project into the future")
            
            curve_fit_params = {}
            if enable_curve_fitting:
                col_curve1, col_curve2 = st.columns(2)
                
                with col_curve1:
                    # Model selection
                    fit_model = st.selectbox(
                        "Curve Fitting Model",
                        options=["exponential", "linear", "polynomial", "logarithmic"],
                        index=0,
                        help="Mathematical model to fit to temperature data"
                    )
                    
                    # Future projection time
                    projection_minutes = st.number_input(
                        "Future Projection (minutes)",
                        min_value=1,
                        max_value=1440,  # 24 hours max
                        value=60,
                        help="How far into the future to project the curve"
                    )
                
                with col_curve2:
                    # Channel selection for curve fitting (up to 4)
                    st.subheader("Select Channels to Fit")
                    
                    # Get available columns by directly processing descriptor row
                    available_channels = []
                    try:
                        # Get descriptor row data
                        descriptor_series = df.iloc[descriptor_row]
                        
                        # Find valid temperature columns using same logic as main processing
                        for idx, desc in enumerate(descriptor_series):
                            desc_str = str(desc).lower()
                            
                            # Skip columns containing exclude substrings
                            exclude_patterns = [pattern.strip().lower() for pattern in exclude_substring.split(',') if pattern.strip()]
                            if any(pattern in desc_str for pattern in exclude_patterns):
                                continue
                            
                            # Skip ID column pattern
                            if id_column_pattern.lower() in desc_str:
                                continue
                            
                            # Add valid temperature channels
                            if legend_source == 'descriptor':
                                channel_name = str(desc)
                            else:
                                channel_name = f"Column_{idx}"
                            
                            if channel_name and channel_name != 'nan':
                                available_channels.append(channel_name)
                        
                        if available_channels:
                            st.success(f"Found {len(available_channels)} temperature channels")
                            
                            # Allow selection of up to 4 channels
                            selected_channels = st.multiselect(
                                "Channels for curve fitting (max 4)",
                                options=available_channels,
                                default=available_channels[:min(2, len(available_channels))],
                                help="Select up to 4 temperature channels to fit curves and project"
                            )
                            
                            if len(selected_channels) > 4:
                                st.warning("Maximum 4 channels can be selected for curve fitting")
                                selected_channels = selected_channels[:4]
                        else:
                            selected_channels = []
                            st.warning("No temperature columns found.")
                            st.info(f"Descriptor row {descriptor_row} checked. Exclude patterns: '{exclude_substring}'. ID pattern: '{id_column_pattern}'")
                            
                    except Exception as e:
                        selected_channels = []
                        st.error(f"Error finding channels: {str(e)}")
                        st.info("Please check your data configuration settings.")
                
                # Store curve fitting parameters
                curve_fit_params = {
                    'enabled': True,
                    'model': fit_model,
                    'projection_minutes': projection_minutes,
                    'selected_channels': selected_channels
                }
            else:
                curve_fit_params = {'enabled': False}
            
            # Data Smoothing/Filtering Settings
            st.header("🔧 Data Smoothing & Filtering")
            
            enable_filtering = st.checkbox("Enable data smoothing/filtering", 
                                         help="Apply smoothing to reduce noise in temperature data")
            
            filter_params = {}
            if enable_filtering:
                col_filter1, col_filter2 = st.columns(2)
                
                with col_filter1:
                    # Filter type selection
                    filter_type = st.selectbox(
                        "Smoothing Method",
                        options=["moving_average", "savgol", "exponential"],
                        index=0,
                        help="Method to smooth temperature data:\n"
                             "• Moving Average: Simple rolling average\n"
                             "• Savitzky-Golay: Polynomial smoothing (preserves peaks)\n"
                             "• Exponential: Exponential weighted moving average"
                    )
                    
                    if filter_type == "moving_average":
                        window_size = st.slider(
                            "Window Size (samples)",
                            min_value=3,
                            max_value=51,
                            value=5,
                            step=2,
                            help="Number of samples to average (larger = smoother)"
                        )
                        filter_params = {
                            'enabled': True,
                            'type': 'moving_average',
                            'window': window_size
                        }
                    
                    elif filter_type == "savgol":
                        window_size = st.slider(
                            "Window Size (samples)",
                            min_value=5,
                            max_value=51,
                            value=11,
                            step=2,
                            help="Window length (must be odd, larger = smoother)"
                        )
                        poly_order = st.slider(
                            "Polynomial Order",
                            min_value=1,
                            max_value=min(5, window_size-1),
                            value=2,
                            help="Polynomial order (lower = smoother, higher = preserves features)"
                        )
                        filter_params = {
                            'enabled': True,
                            'type': 'savgol',
                            'window': window_size,
                            'polyorder': poly_order
                        }
                    
                    elif filter_type == "exponential":
                        alpha = st.slider(
                            "Smoothing Factor (α)",
                            min_value=0.01,
                            max_value=1.0,
                            value=0.1,
                            step=0.01,
                            help="Smoothing factor (lower = smoother, higher = more responsive)"
                        )
                        filter_params = {
                            'enabled': True,
                            'type': 'exponential',
                            'alpha': alpha
                        }
                
                with col_filter2:
                    # Preview of filtering effect
                    st.info(f"**Filter Settings:**\n"
                           f"Method: {filter_type.replace('_', ' ').title()}\n" +
                           (f"Window: {window_size} samples" if filter_type in ['moving_average', 'savgol'] else "") +
                           (f", Polynomial: {poly_order}" if filter_type == 'savgol' else "") +
                           (f"Alpha: {alpha}" if filter_type == 'exponential' else ""))
                    
                    st.warning("⚠️ Filtering will be applied to visualization only. Raw data used for calculations.")
            else:
                filter_params = {'enabled': False}
            
            # Temperature Analysis Settings
            st.header("🌡️ Temperature Analysis Settings")
            
            col_analysis1, col_analysis2 = st.columns(2)
            
            with col_analysis1:
                temp_rating = st.number_input(
                    "Temperature Rating (°C)",
                    min_value=0.0,
                    max_value=200.0,
                    value=55.0,
                    step=5.0,
                    help="Product temperature rating for worst-case calculations"
                )
            
            with col_analysis2:
                st.write("")  # Empty space to maintain layout
            

            

            
            # Process data button
            if st.button("🔄 Process Data and Generate Plot", type="primary"):
                start_time = time.time()  # Initialize timing outside try block
                try:
                    
                    # Process the data
                    plot_data, fig = process_temperature_data(
                        df, descriptor_row, first_data_row, sample_interval_sec,
                        id_column_pattern, exclude_substring, legend_source,
                        tick_interval_min, colormap_name, custom_colors_input,
                        curve_fit_params=curve_fit_params,
                        temp_rating=temp_rating,
                        enable_start_trim=enable_start_trim,
                        start_time_hours=start_time_hours,
                        enable_end_trim=enable_end_trim,
                        end_time_hours=end_time_hours,
                        preserve_original_time=preserve_original_time,
                        filter_params=filter_params
                    )
                    
                    if plot_data is not None:
                        # Display the plot
                        st.pyplot(fig, use_container_width=True)
                        # Persist latest plot context for stable export
                        st.session_state['plot_data_latest'] = plot_data
                        st.session_state['sample_interval_sec'] = sample_interval_sec
                        st.session_state['last_plot_figure'] = fig
                        if hasattr(uploaded_file, 'name'):
                            st.session_state['last_original_filename'] = uploaded_file.name.rsplit('.', 1)[0]
                        

                        
                        # Display Final Data Table
                        if 'final_data_table' in plot_data:
                            st.header("📊 Final Data Table - Worst Case Analysis")
                            final_results = plot_data['final_data_table']
                            
                            # Summary information
                            col_summary1, col_summary2, col_summary3 = st.columns(3)
                            
                            with col_summary1:
                                st.metric("Hottest Temperature", 
                                         f"{final_results['max_temperature']:.2f}°C",
                                         delta=f"at {final_results['hottest_time_hours']:.2f} hours")
                            
                            with col_summary2:
                                st.metric("Hottest Channel", 
                                         final_results['hottest_channel'])
                            
                            with col_summary3:
                                st.metric("Ambient Reference", 
                                         f"{final_results['ambient_temp']:.2f}°C",
                                         delta=f"({final_results['ambient_channel']})")
                            
                            # Final data table
                            st.subheader("Temperature Analysis at Hottest Time")
                            final_df = pd.DataFrame(final_results['table_data'])
                            
                            # Style the dataframe with text colors
                            def color_special_rows(row):
                                if row['Is Hottest']:
                                    return ['color: #d32f2f; font-weight: bold'] * len(row)
                                elif row['Is Ambient']:
                                    return ['color: #388e3c; font-weight: bold'] * len(row)
                                else:
                                    return [''] * len(row)
                            
                            styled_df = final_df.style.apply(color_special_rows, axis=1)
                            st.dataframe(styled_df, use_container_width=True)
                            
                            st.info("🔴 Red Text: Hottest Channel | 🟢 Green Text: Ambient Reference Channel")
                            
                        

                        
                        # Export section moved to bottom of the app. Please use the Export Report panel at the end of the page.
                        st.info("Scroll to the bottom for the Export Report section (plot downloads and Excel export).")
                        
                        # Mark that we rendered the export section in this run
                        st.session_state['export_section_shown'] = True
                        
                        # Processing time for informational display
                        processing_time = time.time() - start_time
                        
                        # Display summary information
                        st.success(f"✅ Plot generated successfully with {len(plot_data['channels'])} temperature channels")
                        
                        # Display Stability Analysis
                        if 'stability_analysis' in plot_data:
                            st.header("🔄 Temperature Stability Analysis")
                            stability_results = plot_data['stability_analysis']
                            
                            st.info("🔍 Stability check: Calculating temperature change over three consecutive 30-sample intervals from the end of the data. A channel is unstable if any change exceeds 1.0°C.")
                            st.info("📊 Sample 1, 2, 3: The three temperature readings used for stability analysis (Sample 1 = latest, Sample 3 = earliest). SAN shown in parentheses for traceability.")
                            st.warning("📋 Sample numbers (#) shown below correspond to actual Excel row numbers (accounting for data row configuration)")
                            
                            # Summary metrics
                            stable_channels = [ch for ch, result in stability_results.items() if result['stable']]
                            unstable_channels = [ch for ch, result in stability_results.items() if not result['stable']]
                            
                            col_stable1, col_stable2, col_stable3 = st.columns(3)
                            with col_stable1:
                                st.metric("Stable Channels", len(stable_channels))
                            with col_stable2:
                                st.metric("Unstable Channels", len(unstable_channels))
                            with col_stable3:
                                total_channels = len(stability_results)
                                stability_rate = (len(stable_channels) / total_channels * 100) if total_channels > 0 else 0
                                st.metric("Stability Rate", f"{stability_rate:.1f}%")
                            
                            # Detailed results table
                            stability_table_data = []
                            for channel, result in stability_results.items():
                                if result['stable']:
                                    status = "✅ Stable"
                                    start_time = f"{result['stability_start_time']:.2f}h" if result.get('stability_start_time') else "N/A"
                                    end_time = f"{result['stability_end_time']:.2f}h" if result.get('stability_end_time') else "N/A"
                                    duration = f"{result['stability_duration']:.2f}h" if result.get('stability_duration', 0) > 0 else "N/A"
                                    avg_temp = f"{result['average_temp']:.2f}°C" if result.get('average_temp') else "N/A"
                                    max_dev = f"{result['max_deviation']:.2f}°C" if result.get('max_deviation') is not None else "N/A"
                                else:
                                    status = "❌ Unstable"
                                    start_time = end_time = duration = "N/A"
                                    avg_temp = f"{result['average_temp']:.2f}°C" if result.get('average_temp') else "N/A"
                                    max_dev = f"{result['max_deviation']:.2f}°C" if result.get('max_deviation') is not None else "N/A"
                                
                                # Extract the 3 sample points with sample numbers (show for both stable and unstable)
                                if result['stability_points'] and len(result['stability_points']) >= 3:
                                    p1 = result['stability_points'][0]
                                    p2 = result['stability_points'][1]
                                    p3 = result['stability_points'][2]
                                    sample1_temp = f"{p1['temp']:.2f}°C"
                                    sample2_temp = f"{p2['temp']:.2f}°C"
                                    sample3_temp = f"{p3['temp']:.2f}°C"
                                    # Prefer SAN; fallback to row/index
                                    san1 = p1.get('san')
                                    san2 = p2.get('san')
                                    san3 = p3.get('san')
                                    ref1 = san1 if san1 is not None and san1 != '' else p1.get('index', 'N/A')
                                    ref2 = san2 if san2 is not None and san2 != '' else p2.get('index', 'N/A')
                                    ref3 = san3 if san3 is not None and san3 != '' else p3.get('index', 'N/A')
                                    sample1 = f"{sample1_temp} ({ref1})" if ref1 != 'N/A' else sample1_temp
                                    sample2 = f"{sample2_temp} ({ref2})" if ref2 != 'N/A' else sample2_temp
                                    sample3 = f"{sample3_temp} ({ref3})" if ref3 != 'N/A' else sample3_temp
                                else:
                                    sample1 = sample2 = sample3 = "N/A"
                                
                                stability_table_data.append({
                                    'Channel': channel,
                                    'Status': status,
                                    'Sample 1': sample1,
                                    'Sample 2': sample2,
                                    'Sample 3': sample3,
                                    'Start Time': start_time,
                                    'End Time': end_time,
                                    'Duration': duration,
                                    'Avg Temperature': avg_temp,
                                    'Max Deviation': max_dev,
                                    'Reason': result['reason']
                                })
                            
                            stability_df = pd.DataFrame(stability_table_data)
                            
                            # Convert temperature columns to numeric for proper sorting
                            def extract_temp_for_sorting(x):
                                if isinstance(x, str) and x != 'N/A' and '°C' in x:
                                    # Extract temperature value before '°C', handling format like "25.30°C (#123)"
                                    temp_part = x.split('°C')[0]
                                    try:
                                        return float(temp_part)
                                    except ValueError:
                                        return float('inf')
                                return float('inf')
                            
                            # Style the stability table
                            def color_stability_rows(row):
                                if "✅ Stable" in row['Status']:
                                    return ['color: #388e3c; font-weight: bold'] * len(row)
                                elif "❌ Unstable" in row['Status']:
                                    return ['color: #d32f2f; font-weight: bold'] * len(row)
                                else:
                                    return [''] * len(row)
                            
                            # Display the table (no sorting controls)
                            styled_display_df = stability_df.style.apply(color_stability_rows, axis=1)
                            st.dataframe(styled_display_df, use_container_width=True)
                            
                            st.info("🟢 Green: Stable channels | 🔴 Red: Unstable channels")
                            
                            # Detailed stability points and data summary removed per user request
                        
                except Exception as e:
                    try:
                        processing_time = time.time() - start_time
                    except:
                        processing_time = 0
                    st.error(f"❌ Error processing data: {str(e)}")
                    st.exception(e)

            # Removed in favor of unified bottom Export Report section
            # (previous fallback export UI deleted)
        
        except Exception as e:
            st.error(f"❌ Error loading Excel file: {str(e)}")
            st.exception(e)
    
    else:
        st.info("👆 Please upload an Excel file to begin analysis")
        
        # Show example of expected data format
        with st.expander("📖 Expected Data Format"):
            st.markdown("""
            **Your Excel file should contain:**
            - A descriptor row with column names/descriptions
            - Temperature measurement data in subsequent rows
            - A time/sample ID column
            - Temperature values in numeric format
            
            **Example structure:**
            ```
            Row 44: [Sample, Sensor_1, Sensor_2, Sensor_3, Alarm_Status]
            Row 45: [1, 25.5, 26.2, 24.8, OK]
            Row 46: [2, 25.7, 26.4, 25.1, OK]
            ...
            ```
            """)

    # Always show a bottom Export Report section if we have plot data
    if 'plot_data_latest' in st.session_state:
        st.divider()
        st.subheader("📄 Export Report")
        st.caption("PDF/SVG downloads export the plot only. Use the Excel export to include trimmed data, the plot image, stability, and metadata.")

        pd_latest = st.session_state['plot_data_latest']
        si_latest = st.session_state.get('sample_interval_sec', st.session_state.get('current_sample_interval', 60))
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Plot-only downloads based on current export format selection
        export_format = st.session_state.get('current_export_format', st.session_state.get('export_format_selection', 'PDF'))
        fig_for_export = pd_latest.get('figure') if isinstance(pd_latest, dict) else None

        # Generate plot-only downloads using the latest figure from the last processing run
        fig_for_export = st.session_state.get('last_plot_figure')
        file_data = b''
        mime_type = 'application/pdf'
        file_extension = 'pdf'
        if fig_for_export is not None:
            if export_format == 'PDF':
                buffer = io.BytesIO()
                fig_for_export.savefig(buffer, format='pdf', bbox_inches='tight', dpi=300)
                buffer.seek(0)
                file_data = buffer.getvalue()
                mime_type = 'application/pdf'
                file_extension = 'pdf'
            elif export_format == 'JPEG':
                buffer = io.BytesIO()
                fig_for_export.savefig(buffer, format='jpeg', bbox_inches='tight', dpi=300, facecolor='white')
                buffer.seek(0)
                file_data = buffer.getvalue()
                mime_type = 'image/jpeg'
                file_extension = 'jpg'
            elif export_format == 'SVG':
                buffer = io.StringIO()
                fig_for_export.savefig(buffer, format='svg', bbox_inches='tight')
                file_data = buffer.getvalue().encode('utf-8')
                mime_type = 'image/svg+xml'
                file_extension = 'svg'
        st.download_button(
            label=f"📥 Download {export_format} (Plot Only)",
            data=file_data,
            file_name=f"temperature_plot_{timestamp}.{file_extension}",
            mime=mime_type,
            use_container_width=True,
            disabled=(fig_for_export is None)
        )
        # Alternate format button
        if fig_for_export is not None:
            alt_label = "📥 Download PDF (Plot Only)" if export_format != 'PDF' else "📥 Download SVG (Plot Only)"
            if export_format != 'PDF':
                pdf_buffer = io.BytesIO()
                fig_for_export.savefig(pdf_buffer, format='pdf', bbox_inches='tight', dpi=300)
                pdf_buffer.seek(0)
                alt_data = pdf_buffer.getvalue()
                alt_name = f"temperature_plot_{timestamp}.pdf"
                alt_mime = 'application/pdf'
            else:
                svg_buffer = io.StringIO()
                fig_for_export.savefig(svg_buffer, format='svg', bbox_inches='tight')
                alt_data = svg_buffer.getvalue().encode('utf-8')
                alt_name = f"temperature_plot_{timestamp}.svg"
                alt_mime = 'image/svg+xml'
            st.download_button(
                label=alt_label,
                data=alt_data,
                file_name=alt_name,
                mime=alt_mime,
                use_container_width=True
            )
        else:
            st.download_button(
                label="📥 Download PDF (Plot Only)",
                data=b"",
                file_name=f"temperature_plot_{timestamp}.pdf",
                mime='application/pdf',
                use_container_width=True,
                disabled=True
            )

        # Prepared Excel quick download if exists
        prepared_bytes = st.session_state.get('prepared_excel_bytes')
        prepared_name = st.session_state.get('prepared_excel_filename', f"trimmed_data_{timestamp}.xlsx")
        prepared_help = st.session_state.get('prepared_excel_help', "Excel export includes trimmed data sheet, a plot sheet at the top, stability results, and your test metadata")
        if prepared_bytes:
            st.success("Excel report prepared. Click below to download.")
            st.download_button(
                label="📊 Download Excel (Trimmed Data + Plot + Stability)",
                data=prepared_bytes,
                file_name=prepared_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                help=prepared_help
            )
            st.caption(f"Filename: {prepared_name}")
            if st.button("Clear prepared Excel", help="Removes the prepared file so you can re-prepare with new settings"):
                st.session_state.pop('prepared_excel_bytes', None)
                st.session_state.pop('prepared_excel_filename', None)
                st.session_state.pop('prepared_excel_help', None)

        # Minimal Excel flow: button first, then metadata form
        show_excel_form = st.session_state.get('show_excel_meta_form', False)
        if not show_excel_form:
            st.button("Prepare Excel", type="primary", use_container_width=True, key="bottom_prepare_excel", on_click=lambda: st.session_state.update({'show_excel_meta_form': True}))
        else:
            with st.form("bottom_excel_export_form"):
                st.markdown("Enter test metadata for Excel export:")
                test_name = st.text_input("Test Name", key="bottom_meta_test_name")
                test_description = st.text_area("Test Description", key="bottom_meta_test_desc", height=80)
                col_meta1, col_meta2 = st.columns(2)
                with col_meta1:
                    test_date = st.text_input("Test Date", key="bottom_meta_test_date", placeholder="YYYY-MM-DD or free text")
                with col_meta2:
                    test_person = st.text_input("Test Person", key="bottom_meta_test_person")
                excel_filename_input = st.text_input("Excel File Name (optional)", key="bottom_meta_excel_filename", placeholder="my_report.xlsx")
                submit_excel = st.form_submit_button("Generate Excel", type="primary")
            if submit_excel:
                try:
                    report_meta = {
                        'test_name': (test_name or '').strip() or None,
                        'test_description': (test_description or '').strip() or None,
                        'test_date': (test_date or '').strip() or None,
                        'test_person': (test_person or '').strip() or None,
                    }
                    original_filename = st.session_state.get('last_original_filename', "temperature_data")
                    excel_buffer = export_trimmed_data_to_excel(
                        pd_latest['time_hours'],
                        pd_latest['data'],
                        si_latest,
                        pd_latest,
                        original_filename,
                        report_meta=report_meta
                    )
                    # Determine prepared filename
                    excel_filename_clean = (excel_filename_input or '').strip()
                    if excel_filename_clean:
                        # Ensure .xlsx extension
                        if not excel_filename_clean.lower().endswith('.xlsx'):
                            excel_filename_clean += '.xlsx'
                        prepared_name = excel_filename_clean
                    else:
                        prepared_name = f"trimmed_data_{timestamp}.xlsx"
                    st.session_state['prepared_excel_bytes'] = excel_buffer.getvalue()
                    st.session_state['prepared_excel_filename'] = prepared_name
                    st.session_state['prepared_excel_help'] = "Excel export includes trimmed data sheet, a plot sheet at the top, stability results, and your test metadata"
                    st.session_state['show_excel_meta_form'] = False
                    st.success("Excel report prepared. Use the download button above.")
                except Exception as e:
                    st.error(f"Excel export failed: {str(e)}")
            if st.button("Cancel", key="bottom_cancel_excel_form", help="Hide metadata form"):
                st.session_state['show_excel_meta_form'] = False


def process_temperature_data(df, descriptor_row, first_data_row, sample_interval_sec,
                           id_column_pattern, exclude_substring, legend_source,
                           tick_interval_min, colormap_name, custom_colors_input,
                           curve_fit_params=None, preview_only=False, 
                           temp_rating=55.0,
                           enable_start_trim=False, start_time_hours=0.0,
                           enable_end_trim=False, end_time_hours=100.0,
                           preserve_original_time=False, filter_params=None):
    """Process temperature data and generate plot"""
    
    try:
        # Validate descriptor row
        descriptor_series = df.iloc[descriptor_row]
        
        # Auto-detect better descriptor row if current one looks wrong
        string_cells = descriptor_series.astype(str).str.contains(r'[a-zA-Z]', na=False).sum()
        total_cells = len(descriptor_series.dropna())
        
        if total_cells > 0 and string_cells / total_cells <= 0.1:
            st.warning("⚠️ Descriptor row might be incorrect. Scanning for better row...")
            for i in range(descriptor_row + 1, min(descriptor_row + 10, df.shape[0])):
                test_series = df.iloc[i]
                test_string_cells = test_series.astype(str).str.contains(r'[a-zA-Z]', na=False).sum()
                test_total_cells = len(test_series.dropna())
                
                if test_total_cells > 0 and test_string_cells / test_total_cells > 0.5:
                    descriptor_row = i
                    descriptor_series = df.iloc[descriptor_row]
                    st.info(f"✅ Found better descriptor row at index {descriptor_row}")
                    break
        
        # Identify columns to keep
        columns_to_keep = []
        id_column = None
        
        for idx, desc in enumerate(descriptor_series):
            desc_str = str(desc).lower()
            header_str = str(df.columns[idx]).lower() if idx < len(df.columns) else ""
            
            # Skip columns containing any exclude substring
            exclude_patterns = [pattern.strip().lower() for pattern in exclude_substring.split(',') if pattern.strip()]
            if any(pattern in desc_str for pattern in exclude_patterns):
                continue
                
            # Identify ID column
            if (id_column_pattern.lower() in desc_str or 
                id_column_pattern.lower() in header_str) and id_column is None:
                id_column = idx
                continue
                
            columns_to_keep.append(idx)
        
        # If no ID column found by pattern, find monotonically increasing integer column
        if id_column is None:
            data_subset = df.iloc[first_data_row:first_data_row+min(10, df.shape[0]-first_data_row)]
            for idx in range(df.shape[1]):
                if idx in columns_to_keep:
                    continue
                try:
                    col_series = data_subset.iloc[:, idx]
                    col_data = pd.to_numeric(col_series, errors='coerce').dropna()
                    if len(col_data) > 1 and col_data.is_monotonic_increasing and col_data.dtype.kind in 'iu':
                        id_column = idx
                        break
                except:
                    continue
        
        if not columns_to_keep:
            raise ValueError("No valid temperature columns found after filtering")
        
        # Extract data
        data = df.iloc[first_data_row:].reset_index(drop=True)
        
        # Convert temperature columns to numeric
        temp_data = {}
        channel_names = []
        
        for col_idx in columns_to_keep:
            # Get column name based on legend source
            if legend_source == 'descriptor':
                col_name = str(descriptor_series.iloc[col_idx])
            else:
                col_name = f"Column_{col_idx}" if col_idx >= len(df.columns) else str(df.columns[col_idx])
            
            # Convert to numeric
            numeric_data = pd.to_numeric(data.iloc[:, col_idx], errors='coerce')
            temp_data[col_name] = numeric_data
            channel_names.append(col_name)
        
        # Create time axis
        time_sec = np.arange(len(data)) * sample_interval_sec
        time_hours = time_sec / 3600
        
        # Sort channels by final temperature (descending)
        final_temps = {}
        for channel in channel_names:
            temp_series = temp_data[channel].dropna()
            if not temp_series.empty:
                final_temps[channel] = temp_series.iloc[-1]
            else:
                final_temps[channel] = -999  # Put channels with no data at the end
        
        sorted_channels = sorted(channel_names, key=lambda x: final_temps[x], reverse=True)
        
        # Trimming: if global trimming is enabled, apply it and skip manual trimming to avoid double-trimming
        if st.session_state.get('global_trim_enabled', False):
            time_hours, temp_data = apply_global_trimming(time_hours, temp_data)
            sorted_channels = [ch for ch in sorted_channels if ch in temp_data and not temp_data[ch].empty]
            st.info("🎯 Global trimming applied to analysis data")
        else:
            # Apply time trimming if enabled (manual)
            trimming_applied = False
            original_length = len(time_hours)
            
            if enable_start_trim and start_time_hours > 0:
                # Find the index where time >= start_time_hours
                start_trim_idx = np.where(time_hours >= start_time_hours)[0]
                
                if len(start_trim_idx) > 0:
                    start_idx = start_trim_idx[0]
                    trimming_applied = True
                else:
                    start_idx = 0
                    st.warning(f"⚠️ Start time {start_time_hours:.2f} hours is beyond data range.")
            else:
                start_idx = 0
            
            if enable_end_trim and end_time_hours < time_hours[-1]:
                # Find the last index where time <= end_time_hours
                end_trim_idx = np.where(time_hours <= end_time_hours)[0]
                
                if len(end_trim_idx) > 0:
                    end_idx = end_trim_idx[-1] + 1  # +1 to include the last valid point
                    trimming_applied = True
                else:
                    end_idx = len(time_hours)
                    st.warning(f"⚠️ End time {end_time_hours:.2f} hours is before data range.")
            else:
                end_idx = len(time_hours)
            
            # Apply trimming if any trimming is enabled and valid
            if trimming_applied and start_idx < end_idx:
                # Trim time axis
                time_hours = time_hours[start_idx:end_idx]
                
                # Reset time to start from 0 when not preserving original time and any trimming applied
                if (not preserve_original_time) and (enable_start_trim or enable_end_trim):
                    # Rebuild a clean 0-based time axis using sample interval to ensure consistent spacing
                    time_hours = np.arange(len(time_hours)) * (sample_interval_sec / 3600.0)
                
                # Trim temperature data for all channels
                trimmed_temp_data = {}
                for channel, temps in temp_data.items():
                    if start_idx < len(temps):
                        trimmed_data = temps.iloc[start_idx:min(end_idx, len(temps))].reset_index(drop=True)
                        trimmed_temp_data[channel] = trimmed_data
                    else:
                        # If start_idx is beyond data length, create empty series
                        trimmed_temp_data[channel] = pd.Series(dtype=float)
                
                temp_data = trimmed_temp_data
                
                # Update sorted channels to only include those with data
                sorted_channels = [ch for ch in sorted_channels if not temp_data[ch].empty]
                
                # Show trimming information
                samples_removed = original_length - len(time_hours)
                trim_info = []
                if enable_start_trim and start_time_hours > 0:
                    trim_info.append(f"before {start_time_hours:.2f}h")
                if enable_end_trim and end_time_hours < (original_length * sample_interval_sec / 3600):
                    trim_info.append(f"after {end_time_hours:.2f}h")
                
                if trim_info:
                    time_axis_msg = " (original time preserved)" if preserve_original_time else " (time reset to 0)"
                    st.info(f"✂️ Data trimmed: Removed {samples_removed} samples {' and '.join(trim_info)}{time_axis_msg}")
            elif trimming_applied:
                st.error("❌ Invalid trimming range: Start time must be before end time")
        
        # Re-sort channels by final temperature of TRIMMED data (hottest at last sample)
        final_temps_trimmed = {}
        for channel in sorted_channels:
            temp_series = temp_data[channel].dropna()
            if not temp_series.empty:
                final_temps_trimmed[channel] = temp_series.iloc[-1]  # Last sample of trimmed data
            else:
                final_temps_trimmed[channel] = -999  # Put channels with no data at the end
        
        # Sort by final temperature of trimmed data (descending - hottest first)
        sorted_channels = sorted(sorted_channels, key=lambda x: final_temps_trimmed[x], reverse=True)
        
        # For preview mode, just return the data structure
        if preview_only:
            return {
                'time_hours': time_hours,
                'data': temp_data,
                'channels': sorted_channels
            }, None
        
        # Handle custom colors
        custom_colors = None
        if custom_colors_input.strip():
            try:
                custom_colors = [color.strip() for color in custom_colors_input.split(',') if color.strip()]
            except:
                st.warning("⚠️ Invalid custom colors format. Using default colormap.")
        
        # Create plot
        fig, ax = plt.subplots(figsize=(12, 8))
        
        # Get colors
        if custom_colors and len(custom_colors) > 0:
            colors = custom_colors * ((len(sorted_channels) // len(custom_colors)) + 1)
        else:
            cmap = plt.get_cmap(colormap_name)
            colors = [cmap(i / max(1, len(sorted_channels) - 1)) for i in range(len(sorted_channels))]
        
        # Apply filtering if enabled (for display only)
        filtered_temp_data = {}
        if filter_params and filter_params.get('enabled', False):
            for channel in sorted_channels:
                temp_series = temp_data[channel].dropna()
                if len(temp_series) < 5:  # Need minimum data for filtering
                    filtered_temp_data[channel] = temp_series
                    continue
                
                try:
                    if filter_params['type'] == 'moving_average':
                        window = min(filter_params['window'], len(temp_series))
                        filtered_series = temp_series.rolling(window=window, center=True).mean()
                        # Fill NaN values at edges
                        filtered_series = filtered_series.bfill().ffill()
                        filtered_temp_data[channel] = filtered_series
                    
                    elif filter_params['type'] == 'savgol':
                        window = min(filter_params['window'], len(temp_series))
                        if window % 2 == 0:  # Ensure odd window
                            window -= 1
                        if window < 5:
                            window = 5
                        polyorder = min(filter_params['polyorder'], window - 1)
                        filtered_values = savgol_filter(temp_series.values, window, polyorder)
                        filtered_temp_data[channel] = pd.Series(filtered_values, index=temp_series.index)
                    
                    elif filter_params['type'] == 'exponential':
                        alpha = filter_params['alpha']
                        filtered_series = temp_series.ewm(alpha=alpha).mean()
                        filtered_temp_data[channel] = filtered_series
                    
                    else:
                        filtered_temp_data[channel] = temp_series
                        
                except Exception as e:
                    st.warning(f"Could not apply filter to {channel}: {str(e)}")
                    filtered_temp_data[channel] = temp_series
        else:
            # No filtering, use original data
            for channel in sorted_channels:
                filtered_temp_data[channel] = temp_data[channel]
        
        # Determine which channels to plot
        channels_to_plot = list(sorted_channels)
        if curve_fit_params and curve_fit_params.get('enabled', False):
            selected_channels = curve_fit_params.get('selected_channels', []) or []
            if len(selected_channels) > 0:
                channels_to_plot = [ch for ch in sorted_channels if ch in selected_channels]
        
        # Plot each channel (using filtered data for display)
        for i, channel in enumerate(channels_to_plot):
            temp_series = filtered_temp_data[channel]
            color = colors[i % len(colors)]
            ax.plot(time_hours[:len(temp_series)], temp_series, label=channel, color=color, linewidth=1.5)
        
        # Add curve fitting if enabled
        curve_fit_results = {}
        if curve_fit_params and curve_fit_params.get('enabled', False) and not preview_only:
            model_type = curve_fit_params.get('model', 'exponential')
            projection_minutes = curve_fit_params.get('projection_minutes', 60)
            selected_channels = curve_fit_params.get('selected_channels', [])
            
            st.info(f"🔄 Applying {model_type} curve fitting to {len(selected_channels)} channels...")
            
            # Fit curves for selected channels
            for channel in selected_channels:
                if channel in temp_data and channel in sorted_channels:
                    # Get channel data
                    channel_temps = temp_data[channel].dropna()
                    if len(channel_temps) < 3:
                        st.warning(f"Not enough data points for curve fitting in channel: {channel}")
                        continue
                    
                    # Get corresponding trimmed time data (align with displayed time axis)
                    channel_time_hours = time_hours[:len(channel_temps)]
                    channel_time_sec = channel_time_hours * 3600
                    
                    # Fit curve
                    params, r2, model_func = fit_temperature_curve(
                        channel_time_sec, channel_temps.values, model_type
                    )
                    
                    if params is not None and model_func is not None:
                        # Store results
                        curve_fit_results[channel] = {
                            'params': params,
                            'r2': r2,
                            'model_func': model_func,
                            'model_type': model_type
                        }
                        
                        # Generate future projection
                        future_time_sec, future_temps = generate_future_projection(
                            params, model_func, channel_time_sec, projection_minutes
                        )
                        future_time_hours = future_time_sec / 3600
                        
                        # Plot fitted curve (overlay across existing data range)
                        fitted_time_hours = channel_time_sec / 3600
                        fitted_temps = model_func(channel_time_sec, *params)
                        
                        # Get channel color
                        # Use plotting index color to match displayed series color
                        try:
                            channel_idx = channels_to_plot.index(channel)
                        except ValueError:
                            channel_idx = sorted_channels.index(channel)
                        channel_color = colors[channel_idx % len(colors)]
                        
                        # Overlay dashed fit over historical data
                        ax.plot(fitted_time_hours, fitted_temps,
                               color=channel_color, linestyle='--', alpha=0.9, linewidth=2)
                        
                        # Plot future projection as dashed line
                        ax.plot(future_time_hours, future_temps, 
                               color=channel_color, linestyle='--', alpha=0.7, linewidth=2,
                               label=f'{channel} (projected R²={r2:.3f})')
                        
                        # Add vertical line to show transition to projection
                        ax.axvline(x=fitted_time_hours[-1], color=channel_color, 
                                  linestyle=':', alpha=0.5, linewidth=1)
            
            # Add a general vertical line to mark the end of actual data if any projections exist
            if curve_fit_results:
                actual_data_end = time_hours[-1] if len(time_hours) > 0 else 0
                ax.axvline(x=actual_data_end, color='gray', linestyle='--', alpha=0.8, linewidth=2)
                ax.text(actual_data_end + 0.1, ax.get_ylim()[1] * 0.95, 'Projection Start', 
                       rotation=90, verticalalignment='top', fontsize=10, alpha=0.7)
            
            # Display curve fitting results
            if curve_fit_results:
                st.success(f"✅ Curve fitting completed for {len(curve_fit_results)} channels")
                
                # Show fitting statistics
                with st.expander("📊 Curve Fitting Results", expanded=True):
                    for channel, results in curve_fit_results.items():
                        col1, col2 = st.columns(2)
                        with col1:
                            st.write(f"**{channel}**")
                            st.write(f"Model: {results['model_type'].title()}")
                            st.write(f"R² Score: {results['r2']:.4f}")
                        
                        with col2:
                            if results['model_type'] == 'exponential':
                                T_inf, T_0, tau = results['params']
                                st.write(f"T∞: {T_inf:.2f}°C")
                                st.write(f"T₀: {T_0:.2f}°C") 
                                st.write(f"τ: {tau:.1f}s")
                            elif results['model_type'] == 'linear':
                                a, b = results['params']
                                st.write(f"Slope: {a:.4f}°C/s")
                                st.write(f"Intercept: {b:.2f}°C")
                            elif results['model_type'] == 'polynomial':
                                a, b, c = results['params']
                                st.write(f"a: {a:.6f}")
                                st.write(f"b: {b:.4f}")
                                st.write(f"c: {c:.2f}")
                            elif results['model_type'] == 'logarithmic':
                                a, b = results['params']
                                st.write(f"a: {a:.4f}")
                                st.write(f"b: {b:.2f}")
        
        # Format plot
        ax.set_xlabel('Time (hours)')
        ax.set_ylabel('Temperature (°C)')
        
        # Update title to include filtering and curve fitting info
        title = f'Temperature Measurements vs. Time ({tick_interval_min} min ticks)\nLegend sorted by final temperature (hottest first)'
        
        # Add filtering info
        if filter_params and filter_params.get('enabled', False):
            filter_type = filter_params['type'].replace('_', ' ').title()
            if filter_params['type'] == 'moving_average':
                title += f'\nSmoothed: {filter_type} (window={filter_params["window"]})'
            elif filter_params['type'] == 'savgol':
                title += f'\nSmoothed: {filter_type} (window={filter_params["window"]}, poly={filter_params["polyorder"]})'
            elif filter_params['type'] == 'exponential':
                title += f'\nSmoothed: {filter_type} (α={filter_params["alpha"]:.2f})'
        
        # Add curve fitting info
        if curve_fit_results:
            current_model = curve_fit_params.get('model', 'exponential')
            current_projection = curve_fit_params.get('projection_minutes', 60)
            title += f'\nDashed lines: {current_model} projections ({current_projection} min)'
        ax.set_title(title)
        
        ax.grid(True, alpha=0.3)
        
        # Format x-axis ticks (extend for projections)
        # Base max time from available time array; if empty, estimate from data length as a fallback
        if len(time_hours) > 0:
            base_max_time = time_hours[-1]
        else:
            # Fallback: largest available series length defines span
            longest_len = 0
            for s in temp_data.values():
                try:
                    longest_len = max(longest_len, len(s.dropna()))
                except Exception:
                    continue
            base_max_time = ((longest_len - 1) * sample_interval_sec / 3600.0) if longest_len > 0 else 1
        if curve_fit_results:
            # Extend x-axis to show projections
            current_projection = curve_fit_params.get('projection_minutes', 60)
            projection_hours = current_projection / 60
            max_time_hours = base_max_time + projection_hours
            
            # Adjust tick interval for extended time range to prevent overcrowding
            total_duration_hours = max_time_hours
            if total_duration_hours > 12:  # For long projections, use larger intervals
                adjusted_tick_interval = max(tick_interval_min, 60)  # At least 1 hour intervals
            elif total_duration_hours > 6:
                adjusted_tick_interval = max(tick_interval_min, 30)  # At least 30 min intervals  
            else:
                adjusted_tick_interval = tick_interval_min
        else:
            max_time_hours = base_max_time
            adjusted_tick_interval = tick_interval_min
            
        tick_interval_hours = adjusted_tick_interval / 60
        
        # Generate tick positions with proper spacing
        tick_positions = np.arange(0, max_time_hours + tick_interval_hours/2, tick_interval_hours)
        
        # Limit number of ticks to prevent overcrowding (max 20 ticks)
        if len(tick_positions) > 20:
            # Reduce number of ticks by increasing interval
            new_interval = max_time_hours / 15  # Target ~15 ticks
            tick_interval_hours = new_interval
            tick_positions = np.arange(0, max_time_hours + tick_interval_hours/2, tick_interval_hours)
        
        ax.set_xticks(tick_positions)
        
        # Format tick labels as H:MM with rotation for better readability
        tick_labels = []
        for hours in tick_positions:
            h = int(hours)
            m = int((hours - h) * 60)
            tick_labels.append(f"{h}:{m:02d}")
        
        ax.set_xticklabels(tick_labels, rotation=45 if len(tick_positions) > 10 else 0, ha='right' if len(tick_positions) > 10 else 'center')
        
        # Optimize x-axis range to ensure projections are visible without resizing
        if len(time_hours) > 0:
            x_start = time_hours[0] if preserve_original_time else 0
            x_end = max_time_hours  # includes projection range when applicable
            time_range = max(x_end - x_start, 1e-6)
            x_margin = time_range * 0.02  # 2% margin on each side
            ax.set_xlim(x_start - x_margin, x_end + x_margin)
        
        # Position legend outside plot area
        ax.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
        
        plt.tight_layout()
        
        # Generate final data table
        final_data_results = generate_final_data_table(time_hours, temp_data, sample_interval_sec, temp_rating)
        
        # Perform stability analysis (starts from end of data)
        # Calculate the row offset to map trimmed indices back to Excel row numbers
        row_offset = first_data_row  # This accounts for the configuration offset
        
        # If trimming was applied, we need to account for that offset too
        if enable_start_trim or enable_end_trim or st.session_state.get('global_trim_enabled', False):
            # Find how many samples were trimmed from the beginning
            original_time_sec = np.arange(len(data)) * sample_interval_sec
            original_time_hours = original_time_sec / 3600
            
            if len(time_hours) > 0 and len(original_time_hours) > 0:
                # Find the offset by comparing first time point
                first_trimmed_time = time_hours[0]
                trim_start_samples = np.argmin(np.abs(original_time_hours - first_trimmed_time))
                row_offset += trim_start_samples
        
        # Attempt to extract SAN/Scan column from the data for reference
        san_series = None
        try:
            # Look for a column whose descriptor or header includes 'san' or 'scan'
            san_col_idx = None
            for idx in range(df.shape[1]):
                desc_str = str(descriptor_series.iloc[idx]).lower() if idx < len(descriptor_series) else ''
                header_str = str(df.columns[idx]).lower() if idx < len(df.columns) else ''
                if ('san' in desc_str or 'scan' in desc_str or 'san' in header_str or 'scan' in header_str):
                    san_col_idx = idx
                    break
            if san_col_idx is not None:
                san_series = df.iloc[first_data_row:, san_col_idx].reset_index(drop=True)
        except Exception:
            san_series = None

        stability_results = analyze_temperature_stability(time_hours, temp_data, sample_interval_sec, row_offset=row_offset, san_series=san_series)
        
        return {
            'time_hours': time_hours,
            'data': temp_data,
            'channels': sorted_channels,
            'final_data_table': final_data_results,
            'stability_analysis': stability_results
        }, fig
        
    except Exception as e:
        st.error(f"Error in data processing: {str(e)}")
        raise e


# Database-dependent dialogs removed per user request


if __name__ == "__main__":
    main()
